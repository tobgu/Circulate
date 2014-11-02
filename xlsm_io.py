from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

FIX_SIGN = '*'


def write_simulation_result(result, source, destination):
    wb = load_workbook(source)
    add_simulation_statistics(result, wb)
    add_table_by_participants(result['conference'], wb)
    add_seatings(result['conference'], wb)
    wb.save(filename=destination)


def add_simulation_statistics(result, wb):
    delete_sheet(wb, "Simulation_Statistics")
    ws = wb.create_sheet()
    ws.title = "Simulation_Statistics"

    set_at_row(ws, 1, "Simulation score", result['score'])
#    set_at_row(ws, 2, "Simulated moves", result['total_tests_count'])
#    set_at_row(ws, 3, "Starting points", result['total_iteration_count'])
#    set_at_row(ws, 4, "Simulation duration", result['duration'])

    ws.cell('A8').value = "Relation"
    ws.cell('B8').value = "At same table"
    ws.cell('C8').value = "Badness"
    row = 9
    for (p1, p2, count, badness) in result['relations']:
        set_triplet_at_row(ws, row, '%s - %s' % (result['conference']['staff_names'][p1],
                                                 result['conference']['staff_names'][p2]),
                                                 count, badness)
        row += 1

    ws.column_dimensions['A'].width = 30.0
    ws.column_dimensions['B'].width = 30.0
    ws.column_dimensions['C'].width = 30.0


def straight_table_list(staff_names, tables):
    participants = []
    for i, table in enumerate(tables):
        for p in table:
            participants.append((staff_names[p['id']], i))

    return sorted(participants, key=lambda x: x[0])


def delete_sheet(wb, name):
    try:
        ws = wb[name]
        wb.remove_sheet(ws)
    except KeyError:
        pass


def set_bold(cell):
    cell.style = cell.style.copy(font=cell.style.font.copy(bold=True))


def add_seatings(conference, wb):
    delete_sheet(wb, "Seating_Results")

    ws = wb.create_sheet(0)
    ws.title = "Seating_Results"

    for col_idx, occasion in enumerate(conference['placements']):
        name = occasion['name']
        tables = occasion['tables']
        col = get_column_letter(col_idx+1)
        row = 1
        ws.cell('%s%s' % (col, row)).value = name
        set_bold(ws.cell('%s%s' % (col, row)))
        for i, table in enumerate(tables):
            row += 2
            ws.cell('%s%s' % (col, row)).value = "Table %s" % (i+1)
            set_bold(ws.cell('%s%s' % (col, row)))
            for name, fix in sorted([(conference['staff_names'][p['id']], p['fix']) for p in table]):
                row += 1
                value = "%s%s" % (FIX_SIGN if fix else '', name)
                ws.cell('%s%s' % (col, row)).value = value

        ws.column_dimensions[col].width = 20.0


def add_table_by_participants(conference, wb):
    delete_sheet(wb, "Participant_Results")
    ws = wb.create_sheet(0)
    ws.title = "Participant_Results"

    col_ix = 1
    for occasion in conference['placements']:
        name = occasion['name']
        tables = occasion['tables']
        col = get_column_letter(col_ix)
        col_table = get_column_letter(col_ix+1)
        row = 1
        ws.cell('%s%s' % (col, row)).value = name
        set_bold(ws.cell('%s%s' % (col, row)))

        for participant, table_no in straight_table_list(conference['staff_names'], tables):
            row += 1
            ws.cell('%s%s' % (col, row)).value = participant
            ws.cell('%s%s' % (col_table, row)).value = table_no + 1

        ws.column_dimensions[col].width = 20.0
        ws.column_dimensions[col_table].width = 5
        col_ix += 3


def set_at_row(ws, row, name, value):
    ws.cell('A%s' % row).value = name
    set_bold(ws.cell('A%s' % row))
    ws.cell('B%s' % row).value = value


def set_triplet_at_row(ws, row, a, b, c):
    ws.cell('A%s' % row).value = a
    ws.cell('B%s' % row).value = b
    ws.cell('C%s' % row).value = c


def read_conference_data(filename='seating.xlsm'):
    wb = load_workbook(filename=filename)
    conference = {}
    conference['staff_names'], conference['weight_matrix'], conference['group_names'], conference['group_participation'] = staff_info(wb)
    conference['seating_names'], conference['table_sizes'] = tables_sizes_per_seating(wb)
    conference['guests'] = guests_per_seating(wb)
    adjust_seating(conference, wb)
    return conference


def tables_sizes_per_seating(wb):
    seatings = wb['Seatings']
    seating_names = [r[0].value for r in seatings.rows[2:] if r[0].value]
    table_sizes = [[c.value for c in r[4:] if c.value] for r in seatings.rows[2:]]
    table_sizes = [s for s in table_sizes if s]
    return seating_names, table_sizes


def staff_info(wb):
    staff = wb['Staff']
    group_names = [c.value for c in staff.rows[0][1:]]
    group_weights = [c.value for c in staff.rows[1][1:]]
    participant_names = [r[0].value.strip() for r in staff.rows[2:]]
    group_participation_weights = [[int(c.value) * group_weights[i] if c.value else 0 for i, c in enumerate(r[1:])] for r in staff.rows[2:]]
    group_participation = [[i for i, c in enumerate(r[1:]) if c.value is not None] for r in staff.rows[2:]]
    weight_matrix = [[sum(gj * gi for gj, gi in zip(group_participation_weights[j], group_participation_weights[i]))
                      for j in range(len(participant_names))] for i in range(len(participant_names))]

    return participant_names, weight_matrix, group_names, group_participation


def guests_per_seating(wb):
    guests = wb['Guests']
    seatings = [[{'id': i, 'fix': False} for i, r in enumerate(c[2:]) if r.value] for c in guests.columns[1:]]
    return seatings


class InputDataInconsistencyException(Exception):
    pass


def adjust_seating(conference, wb):
    def is_fix(name):
        return name.strip().startswith(FIX_SIGN)

    def strip_fix_sign(name):
        return name if not name.startswith(FIX_SIGN) else name[1:]

    # Use previous optimization result as a starting point if such exists
    try:
        seatings = wb['Seating_Results']
    except KeyError:
        return

    name_to_id = {name: id for id, name in enumerate(conference['staff_names'])}
    try:
        new_seatings = [[{'id': name_to_id[strip_fix_sign(row.value)], 'fix': is_fix(row.value)}
                         for row in col if row.value and not row.style.font.bold]
                        for col in seatings.columns]
    except KeyError as e:
        raise InputDataInconsistencyException(u'Unknown participant %s in previous results' % e)

    for ix, (old, new) in enumerate(zip(conference['guests'], new_seatings)):
        if set(p['id'] for p in old) != set(p['id'] for p in new):
            raise InputDataInconsistencyException(u'Inconsistency in number of participants for occasion %s' % conference['seating_names'][ix])

    conference['guests'] = new_seatings